from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import json
from pathlib import Path
from typing import List, Dict, Any
import io
from generators import (
    generate_full_json, generate_deposant, generate_heritier, 
    generate_compte, generate_contact, generate_representant_legal,
    calculate_parts, generate_idscv, generate_cnie, generate_rib,
    generate_carte, generate_phone, generate_email, generate_date,
    empty_rep
)

app = FastAPI(
    title="SCV Generator API",
    description="API pour générer et valider des données SCV avec cohérence garantie + Import Excel",
    version="2.1.0"
)

# CORS pour Next.js
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3001", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Charger les référentiels
DATA_DIR = Path(__file__).parent / "data"

try:
    with open(DATA_DIR / "ref_corrected.json", 'r', encoding='utf-8') as f:
        refs = json.load(f)['referentiels']
    print("✅ Référentiels chargés depuis ref_corrected.json")
except Exception as e:
    print(f"⚠️  Erreur chargement référentiels: {e}")
    print("⚠️  Utilisation référentiels par défaut")
    # Référentiels par défaut si le fichier n'existe pas
    refs = {
        'villes': [
            {'ville': 'Casablanca', 'code': 'CAS'},
            {'ville': 'Rabat', 'code': 'RAB'},
            {'ville': 'Marrakech', 'code': 'MAR'},
            {'ville': 'Fes', 'code': 'FES'},
            {'ville': 'Tanger', 'code': 'TAN'}
        ]
    }

@app.get("/")
def read_root():
    return {
        "message": "SCV Generator API v2.1 - Avec variabilité, cohérence et import Excel",
        "version": "2.1.0",
        "features": [
            "✅ Variabilité: PP/PM, vivant/décédé, 0-8 héritiers, 1-3 comptes",
            "✅ Cohérence: Parts = 10000, nombreCotitulaires auto, validation intégrée",
            "✅ API endpoints: /generate/* pour génération, /validate/* pour validation",
            "✅ Import Excel: Upload fichier Excel pour génération en masse",
            "✅ Template Excel: Téléchargement template pré-rempli"
        ],
        "endpoints": {
            "generation": [
                "POST /generate/full - Génère un SCV complet",
                "POST /generate/deposant - Génère un déposant avec contact",
                "POST /generate/heritier - Génère un héritier (avec index et total pour parts)",
                "POST /generate/compte - Génère un compte COL ou IND",
                "POST /generate/representant - Génère un représentant légal"
            ],
            "validation": [
                "POST /validate/coherence - Valide la cohérence d'un SCV"
            ],
            "excel": [
                "POST /import/excel - Importe un fichier Excel et génère plusieurs SCV",
                "GET /download/template - Télécharge un template Excel avec exemples"
            ],
            "refs": [
                "GET /refs - Récupère les référentiels (villes)"
            ]
        }
    }

@app.get("/refs")
def get_refs():
    """Retourne les référentiels"""
    return refs

@app.post("/generate/full")
def generate_full(count: int = 1):
    """
    Génère un ou plusieurs SCV complets avec cohérence garantie
    
    Args:
        count: Nombre de SCV à générer (1-100)
    
    Returns:
        Si count=1: Un seul SCV
        Si count>1: Liste de SCV
    """
    if count < 1 or count > 100:
        raise HTTPException(status_code=400, detail="count doit être entre 1 et 100")
    
    try:
        results = []
        for _ in range(count):
            scv_data = generate_full_json(refs['villes'])
            results.append(scv_data)
        
        return {
            "count": len(results),
            "data": results if count > 1 else results[0]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/generate/deposant")
def api_generate_deposant():
    """Génère un déposant avec son contact"""
    try:
        deposant_data = generate_deposant(refs['villes'])
        contact_data = generate_contact(refs['villes'])
        
        return {
            "identifiantDeposant": deposant_data,
            "infosContact": contact_data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/generate/heritier")
def api_generate_heritier(index: int = 0, total: int = 4):
    """
    Génère un héritier avec calcul automatique de la part
    
    Args:
        index: Position de l'héritier (0-based)
        total: Nombre total d'héritiers
    
    Returns:
        Héritier avec partHeritage correcte
    """
    try:
        # Générer l'héritier avec index et total
        heritier_data = generate_heritier(index, total, refs['villes'])
        
        return heritier_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/generate/compte")
def api_generate_compte(deposant_id: str = None, index: int = 0):
    """Génère un compte bancaire COL ou IND"""
    try:
        compte_data = generate_compte(deposant_id or generate_idscv(), index)
        return compte_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/generate/representant")
def api_generate_representant():
    """Génère un représentant légal"""
    try:
        return generate_representant_legal(refs['villes'])
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/validate/coherence")
def validate_coherence(data: dict):
    """
    Valide la cohérence d'un SCV complet
    
    Vérifie:
    - Somme parts héritiers = 10000
    - Somme parts cotitulaires = 10000 (pour comptes COL)
    - nombreCotitulaires = len(cotitulaire) + 1
    - nombreHeritiers = len(heritier)
    - nombreComptes = len(compte)
    - isCarteBancaire cohérent avec infosCarteBancaire
    """
    errors = []
    
    try:
        scv = data.get('SCV', data)
        
        # 1. Vérifier héritiers
        heritiers = scv.get('heritier', [])
        if heritiers:
            somme_parts_heritiers = sum(h['identifiantHeritier']['partHeritage'] for h in heritiers)
            if somme_parts_heritiers != 10000:
                errors.append({
                    "field": "heritier",
                    "message": f"Somme parts héritiers = {somme_parts_heritiers} (attendu: 10000)"
                })
            
            nombre_heritiers = scv['identifiantDeposant'].get('nombreHeritiers', 0)
            if len(heritiers) != nombre_heritiers:
                errors.append({
                    "field": "nombreHeritiers",
                    "message": f"nombreHeritiers = {nombre_heritiers} mais {len(heritiers)} héritiers dans la liste"
                })
        
        # 2. Vérifier comptes
        comptes = scv.get('compte', [])
        nombre_comptes = scv['identifiantDeposant'].get('nombreComptes', 0)
        if len(comptes) != nombre_comptes:
            errors.append({
                "field": "nombreComptes",
                "message": f"nombreComptes = {nombre_comptes} mais {len(comptes)} comptes dans la liste"
            })
        
        # 3. Vérifier chaque compte
        for idx, compte in enumerate(comptes):
            infos = compte['infosCompteBancaire']
            nature = infos['natureCompte']
            
            # Vérifier cotitulaires pour COL
            if nature == 'COL':
                cotitulaires = compte.get('cotitulaire', [])
                nombre_cotit = infos.get('nombreCotitulaires')
                
                # Vérifier nombreCotitulaires
                expected_nombre = len(cotitulaires) + 1
                if nombre_cotit != expected_nombre:
                    errors.append({
                        "field": f"compte[{idx}].nombreCotitulaires",
                        "message": f"nombreCotitulaires = {nombre_cotit} mais devrait être {expected_nombre} ({len(cotitulaires)} cotit + 1 tit)"
                    })
                
                # Vérifier somme parts cotitulaires
                if cotitulaires:
                    somme_parts = sum(c['partCoTitulaire'] for c in cotitulaires)
                    if somme_parts != 10000:
                        errors.append({
                            "field": f"compte[{idx}].cotitulaire",
                            "message": f"Somme parts cotitulaires = {somme_parts} (attendu: 10000)"
                        })
            
            # Vérifier cartes bancaires
            is_carte = infos.get('isCarteBancaire', 'N')
            cartes = compte.get('infosCarteBancaire', [])
            
            if is_carte == 'O' and len(cartes) == 0:
                errors.append({
                    "field": f"compte[{idx}].isCarteBancaire",
                    "message": "isCarteBancaire='O' mais aucune carte dans infosCarteBancaire"
                })
            elif is_carte == 'N' and len(cartes) > 0:
                errors.append({
                    "field": f"compte[{idx}].isCarteBancaire",
                    "message": "isCarteBancaire='N' mais des cartes présentes dans infosCarteBancaire"
                })
        
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "message": "✅ SCV cohérent" if len(errors) == 0 else f"⚠️ {len(errors)} erreur(s) détectée(s)"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# NOUVEAUX ENDPOINTS - IMPORT EXCEL
# ============================================================================

@app.post("/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """
    Importe un fichier Excel et génère plusieurs SCV
    
    Format Excel attendu:
    - Onglet "Deposants" : Info déposant (nom, prenom, type, etc.)
    - Onglet "Heritiers" : Héritiers par deposant_id
    - Onglet "Comptes" : Comptes par deposant_id
    
    Returns:
        Liste de SCV générés avec validation
    """
    try:
        from openpyxl import load_workbook
        
        # Vérifier le type de fichier
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Fichier doit être .xlsx ou .xls")
        
        # Lire le fichier Excel
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        
        # Vérifier les onglets requis
        required_sheets = ['Deposants', 'Heritiers', 'Comptes']
        missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
        if missing_sheets:
            raise HTTPException(
                status_code=400, 
                detail=f"Onglets manquants: {', '.join(missing_sheets)}"
            )
        
        # Parser les données
        deposants = parse_deposants_sheet(wb['Deposants'])
        heritiers_map = parse_heritiers_sheet(wb['Heritiers'])
        comptes_map = parse_comptes_sheet(wb['Comptes'])
        
        # Générer les SCV
        scv_list = []
        errors = []
        
        for idx, deposant_row in enumerate(deposants):
            try:
                deposant_id = deposant_row.get('deposant_id')
                
                # Créer le déposant
                deposant = create_deposant_from_row(deposant_row)
                
                # Créer le contact
                contact = create_contact_from_row(deposant_row, refs['villes'])
                
                # Récupérer les héritiers
                heritiers_rows = heritiers_map.get(deposant_id, [])
                heritiers = []
                for h_row in heritiers_rows:
                    heritiers.append(create_heritier_from_row(h_row, refs['villes']))
                
                # Vérifier cohérence héritiers
                if len(heritiers) != deposant['nombreHeritiers']:
                    deposant['nombreHeritiers'] = len(heritiers)
                
                # Recalculer les parts héritiers
                if heritiers:
                    parts = calculate_parts(len(heritiers))
                    for i, h in enumerate(heritiers):
                        h['identifiantHeritier']['partHeritage'] = parts[i]
                
                # Récupérer les comptes
                comptes_rows = comptes_map.get(deposant_id, [])
                comptes = []
                for c_row in comptes_rows:
                    comptes.append(create_compte_from_row(c_row, deposant['idscv']))
                
                # Vérifier cohérence comptes
                if len(comptes) != deposant['nombreComptes']:
                    deposant['nombreComptes'] = len(comptes)
                
                # Représentant légal (optionnel)
                if deposant_row.get('has_representant', 'N') == 'O':
                    representants = [generate_representant_legal(refs['villes'])]
                else:
                    representants = [empty_rep()]
                
                # Construire le SCV
                scv = {
                    "SCV": {
                        "identifiantDeposant": deposant,
                        "infosContact": contact,
                        "representantLegal": representants,
                        "heritier": heritiers,
                        "compte": comptes
                    }
                }
                
                # Valider
                validation = validate_coherence(scv)
                if not validation['valid']:
                    errors.append({
                        "row": idx + 2,  # +2 car ligne 1 = header
                        "deposant_id": deposant_id,
                        "errors": validation['errors']
                    })
                else:
                    scv_list.append(scv)
                
            except Exception as e:
                errors.append({
                    "row": idx + 2,
                    "deposant_id": deposant_row.get('deposant_id', 'unknown'),
                    "error": str(e)
                })
        
        return {
            "success": True,
            "total_rows": len(deposants),
            "generated": len(scv_list),
            "errors": errors,
            "data": scv_list
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors du traitement: {str(e)}")


@app.get("/download/template")
def download_template():
    """Télécharge un template Excel vide avec exemples"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        # Onglet Deposants
        ws_dep = wb.active
        ws_dep.title = "Deposants"
        
        headers_dep = [
            'deposant_id', 'typePersonne', 'nom', 'prenom', 'dateNaissance',
            'nationalite', 'formeJuridique', 'natureIdentifiant', 'numeroIdentifiant',
            'nombreComptes', 'isDecede', 'nombreHeritiers', 'has_representant',
            'adresse', 'ville', 'mobile', 'fixe', 'email', 'version'
        ]
        
        for idx, header in enumerate(headers_dep, 1):
            cell = ws_dep.cell(1, idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Exemples
        ws_dep.append(['DEP001', 'PP', 'Bourkadi', 'Omar', '15/03/1995', 'MA', None, 'CNIE', 'AB123456', 2, 'O', 3, 'N', '123 Rue Hassan II', 'CAS', '0612345678', None, 'omar@test.ma', 1])
        ws_dep.append(['DEP002', 'PM', 'IMPORT EXPORT SARL', None, None, 'MA', 'SARL', 'RC', 'RC456789', 1, 'N', 0, 'N', '456 Bd Zerktouni', 'CAS', '0623456789', None, 'contact@import.ma', 1])
        
        # Onglet Heritiers
        ws_her = wb.create_sheet("Heritiers")
        
        headers_her = [
            'deposant_id', 'nom', 'prenom', 'dateNaissance', 'nationalite',
            'natureIdentifiant', 'numeroIdentifiant', 'adresse', 'ville', 'mobile', 'email'
        ]
        
        for idx, header in enumerate(headers_her, 1):
            cell = ws_her.cell(1, idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Exemples
        ws_her.append(['DEP001', 'Bourkadi', 'Amine', '20/05/2000', 'MA', 'CNIE', 'CD789012', '789 Rue', 'CAS', '0634567890', 'amine@test.ma'])
        ws_her.append(['DEP001', 'Bourkadi', 'Fatima', '15/08/1998', 'MA', 'CNIE', 'EF345678', '012 Ave', 'CAS', '0645678901', 'fatima@test.ma'])
        ws_her.append(['DEP001', 'Bourkadi', 'Karim', '10/12/2002', 'MA', 'CNIE', 'GH901234', '345 Bd', 'CAS', '0656789012', 'karim@test.ma'])
        
        # Onglet Comptes
        ws_cpt = wb.create_sheet("Comptes")
        
        headers_cpt = [
            'deposant_id', 'rib', 'natureCompte', 'nombreCotitulaires', 'nomCompte',
            'devise', 'isCarteBancaire', 'statutCompte', 'sensSolde',
            'montantSolde', 'montantDettes', 'montantDebits'
        ]
        
        for idx, header in enumerate(headers_cpt, 1):
            cell = ws_cpt.cell(1, idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Exemples
        ws_cpt.append(['DEP001', '007123456789012345678901', 'COL', 3, 'Compte Joint', 'MAD', 'O', 'STP', 'CRED', 2000000000, 50000000, 1000000000])
        ws_cpt.append(['DEP001', '007234567890123456789012', 'IND', None, 'Compte Personnel', 'MAD', 'N', 'STP', 'CRED', 1500000000, 30000000, 800000000])
        ws_cpt.append(['DEP002', '007345678901234567890123', 'IND', None, 'Compte Société', 'MAD', 'O', 'STP', 'CRED', 5000000000, 100000000, 2000000000])
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template_scv.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# FONCTIONS HELPER POUR PARSING EXCEL
# ============================================================================

def parse_deposants_sheet(sheet) -> List[Dict]:
    """Parse l'onglet Deposants"""
    deposants = []
    headers = [cell.value for cell in sheet[1]]
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        deposant = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                deposant[header] = row[idx]
        
        deposants.append(deposant)
    
    return deposants


def parse_heritiers_sheet(sheet) -> Dict[str, List[Dict]]:
    """Parse l'onglet Heritiers et groupe par deposant_id"""
    heritiers_map = {}
    headers = [cell.value for cell in sheet[1]]
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        heritier = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                heritier[header] = row[idx]
        
        deposant_id = heritier.get('deposant_id')
        if deposant_id:
            if deposant_id not in heritiers_map:
                heritiers_map[deposant_id] = []
            heritiers_map[deposant_id].append(heritier)
    
    return heritiers_map


def parse_comptes_sheet(sheet) -> Dict[str, List[Dict]]:
    """Parse l'onglet Comptes et groupe par deposant_id"""
    comptes_map = {}
    headers = [cell.value for cell in sheet[1]]
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        compte = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                compte[header] = row[idx]
        
        deposant_id = compte.get('deposant_id')
        if deposant_id:
            if deposant_id not in comptes_map:
                comptes_map[deposant_id] = []
            comptes_map[deposant_id].append(compte)
    
    return comptes_map


def create_deposant_from_row(row: Dict) -> Dict:
    """Crée un déposant depuis une ligne Excel"""
    type_personne = row.get('typePersonne', 'PP')
    
    return {
        "idscv": row.get('idscv') or generate_idscv(),
        "version": int(row.get('version', 1)),
        "typePersonne": type_personne,
        "nom": row.get('nom'),
        "prenom": row.get('prenom') if type_personne == 'PP' else None,
        "dateNaissance": row.get('dateNaissance') if type_personne == 'PP' else None,
        "nationalite": row.get('nationalite', 'MA'),
        "formeJuridique": row.get('formeJuridique') if type_personne == 'PM' else None,
        "natureIdentifiantDeposant": row.get('natureIdentifiant', 'CNIE'),
        "numeroIdentifiantDeposant": row.get('numeroIdentifiant') or generate_cnie(),
        "denominationSociale": row.get('nom') if type_personne == 'PM' else None,
        "nombreComptes": int(row.get('nombreComptes', 1)),
        "isDecede": row.get('isDecede', 'N'),
        "nombreHeritiers": int(row.get('nombreHeritiers', 0))
    }


def create_contact_from_row(row: Dict, villes: List) -> Dict:
    """Crée un contact depuis une ligne Excel"""
    ville_code = row.get('ville', 'CAS')
    
    return {
        "adresse1": row.get('adresse'),
        "adresse2": None,
        "codePostal": None,
        "ville": ville_code,
        "pays": "MA",
        "mobile": row.get('mobile') or generate_phone(True),
        "fixe": row.get('fixe'),
        "email": row.get('email') or generate_email(row.get('nom', 'test'), row.get('prenom', 'test'))
    }


def create_heritier_from_row(row: Dict, villes: List) -> Dict:
    """Crée un héritier depuis une ligne Excel"""
    nom = row.get('nom')
    prenom = row.get('prenom')
    
    return {
        "identifiantHeritier": {
            "idscv": row.get('idscv') or generate_idscv(),
            "natureIdentifiantDeposantP": row.get('natureIdentifiant', 'CNIE'),
            "numeroIdentifiantDeposantP": row.get('numeroIdentifiant') or generate_cnie(),
            "nom": nom,
            "prenom": prenom,
            "dateNaissance": row.get('dateNaissance') or generate_date(1970, 2000),
            "nationalite": row.get('nationalite', 'MA'),
            "partHeritage": 0  # Sera recalculé automatiquement
        },
        "infosContact": create_contact_from_row(row, villes),
        "representantLegal": empty_rep()
    }


def create_compte_from_row(row: Dict, deposant_id: str) -> Dict:
    """Crée un compte depuis une ligne Excel"""
    nature_compte = row.get('natureCompte', 'IND')
    is_carte = row.get('isCarteBancaire', 'N')
    
    # Cotitulaires pour compte COL
    if nature_compte == 'COL':
        nb_cotitulaires = int(row.get('nombreCotitulaires', 2))
        nb_cotitulaires_liste = nb_cotitulaires - 1
        parts = calculate_parts(nb_cotitulaires_liste)
        
        cotitulaires = []
        for i in range(nb_cotitulaires_liste):
            cotitulaires.append({
                "idscv": generate_idscv(),
                "partCoTitulaire": parts[i]
            })
    else:
        nb_cotitulaires = None
        cotitulaires = []
    
    # Cartes bancaires
    if is_carte == 'O':
        cartes = [{
            "numeroCarte": generate_carte(),
            "validite": "04/2031"
        }]
    else:
        cartes = []
    
    return {
        "infosCompteBancaire": {
            "rib": row.get('rib') or generate_rib(),
            "natureCompte": nature_compte,
            "nombreCotitulaires": nb_cotitulaires,
            "pcec": "2038",
            "devise": row.get('devise', 'MAD'),
            "nomCompte": row.get('nomCompte', f'Compte {nature_compte}'),
            "isCarteBancaire": is_carte,
            "statutCompte": row.get('statutCompte', 'STP'),
            "compteNSTP": None,
            "autreCompteNSTP": None,
            "sensSolde": row.get('sensSolde', 'CRED'),
            "montantTotalSolde": int(row.get('montantSolde', 1500000000)),
            "montantTotalDettes": int(row.get('montantDettes', 50000000)),
            "montantTotalDebits": int(row.get('montantDebits', 1000000000)),
            "montantTotalAgios": 500000000,
            "montantTotalGarantie": 400000000,
            "montantTotalInterets": 1000000000
        },
        "prelevement": [],
        "cotitulaire": cotitulaires,
        "infosCarteBancaire": cartes
    }


if __name__ == "__main__":
    import uvicorn
    print("🚀 Démarrage du serveur SCV Generator API v2.1")
    print("📊 Fonctionnalités:")
    print("   ✅ Génération SCV avec variabilité")
    print("   ✅ Validation cohérence automatique")
    print("   ✅ Import Excel pour génération en masse")
    print("   ✅ Template Excel téléchargeable")
    print("\n📡 Serveur: http://localhost:8001")
    print("📖 Documentation: http://localhost:8001/docs")
    uvicorn.run(app, host="0.0.0.0", port=8001)