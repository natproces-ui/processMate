"""
Générateur Excel universel pour le module d'analyse ProcessMate.
Deux templates : impact (réglementaire/général) et checklist (conformité/couverture).
"""
from __future__ import annotations
import io
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Palette (FF prefix = opaque) ────────────────────────────
BLUE_DARK  = "FF1E3A5F"
BLUE_MID   = "FF2D6A9F"
BLUE_LIGHT = "FFD9EAF7"
GREY_BG    = "FFF8F9FA"
WHITE      = "FFFFFFFF"

COVERAGE_COLORS = {
    "couvert":        ("FFE8F8F5", "FF1E8449"),
    "partiel":        ("FFFFF3CD", "FF9A6A00"),
    "manquant":       ("FFFDECEA", "FFC0392B"),
    "non_applicable": ("FFF8F9FA", "FF6C757D"),
}
COVERAGE_LABELS = {
    "couvert": "✓ Couvert", "partiel": "~ Partiel",
    "manquant": "✗ Manquant", "non_applicable": "— N/A",
}
CRIT_COLORS = {
    "critical": ("FFFDECEA", "FFC0392B"), "high": ("FFFEF3E2", "FFD35400"),
    "medium":   ("FFFFF8E1", "FFB7800A"), "low":  ("FFF0F4F8", "FF546E7A"),
}
CRIT_LABELS = {"critical": "Critique", "high": "Élevé", "medium": "Moyen", "low": "Faible"}

def _font(bold=False, size=10, color="FF000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _fill(hex8): return PatternFill(fill_type="solid", fgColor=hex8)

def _border(style="thin", color="FFD0D7DE"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def _row_height(texts, col_widths):
    mx = 1
    for text, w in zip(texts, col_widths):
        t = str(text or "")
        lines = t.count("\n") + 1
        wrap_est = max(1, len(t) // max(w, 10))
        mx = max(mx, lines, wrap_est)
    return min(max(mx * 14, 18), 220)

def _write_title(ws, title, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = title; c.font = _font(bold=True, size=13, color="FFFFFFFF")
    c.fill = _fill(BLUE_DARK); c.alignment = _align(wrap=False, h="center", v="center")
    ws.row_dimensions[1].height = 30

def _write_headers(ws, col_defs, row=2):
    ws.row_dimensions[row].height = 30
    for ci, (label, width) in enumerate(col_defs, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font = _font(bold=True, size=10, color="FF1E3A5F")
        c.fill = _fill(BLUE_LIGHT); c.alignment = _align(wrap=True, h="center", v="center")
        c.border = _border("medium", "FF4A90D9")
        ws.column_dimensions[get_column_letter(ci)].width = width

# ─── Template Impact ──────────────────────────────────────────

IMPACT_COLS = [
    ("Activité / Procédure", 28), ("Élément source", 34), ("Réf. source", 18),
    ("Impact métier", 52), ("Impact SI", 52), ("Couverture", 14),
    ("Écart identifié", 48), ("Actions recommandées", 60), ("Systèmes", 22),
    ("Section procédure", 30), ("Criticité", 12), ("Dép. externe", 14), ("Confiance", 10),
]

def _write_impact_sheet(ws, analysis, title_prefix):
    ws.title = "Analyse"
    ws.sheet_view.showGridLines = False
    _write_title(ws, f"{title_prefix} — ProcessMate", len(IMPACT_COLS))
    _write_headers(ws, IMPACT_COLS)

    for ri, item in enumerate(analysis, start=3):
        crit = item.get("criticality", "medium")
        cov  = item.get("coverage_status", "manquant")
        crit_bg, crit_fg = CRIT_COLORS.get(crit, CRIT_COLORS["medium"])
        cov_bg, cov_fg   = COVERAGE_COLORS.get(cov, COVERAGE_COLORS["manquant"])
        row_bg = WHITE if ri % 2 == 1 else GREY_BG

        actions_text = "\n".join(
            f"• [{a.get('priority','').upper()}] {a.get('title','')} — {a.get('description','')}"
            for a in (item.get("recommended_actions") or [])
        )
        proc_label = f"{item.get('procedure_ref','') or ''} {item.get('procedure_nom','') or ''}".strip()
        row_data = [
            proc_label,
            item.get("source_element", ""),
            item.get("source_ref", ""),
            item.get("business_impact", ""),
            item.get("si_impact", ""),
            COVERAGE_LABELS.get(cov, cov),
            item.get("gap") or "",
            actions_text,
            "\n".join(item.get("impacted_systems") or []),
            item.get("procedure_section", ""),
            CRIT_LABELS.get(crit, crit),
            item.get("external_dependency") or "",
            f"{round((item.get('confidence') or 0) * 100)}%",
        ]
        ws.row_dimensions[ri].height = _row_height(row_data, [c[1] for c in IMPACT_COLS])

        for ci, value in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.border = _border()
            if ci == 6:  # Couverture
                c.font = _font(bold=True, size=9, color=cov_fg)
                c.fill = _fill(cov_bg); c.alignment = _align(wrap=False, h="center", v="center")
            elif ci == 11:  # Criticité
                c.font = _font(bold=True, size=9, color=crit_fg)
                c.fill = _fill(crit_bg); c.alignment = _align(wrap=False, h="center", v="center")
            else:
                c.font = _font(size=10); c.fill = _fill(row_bg); c.alignment = _align()

    ws.freeze_panes = "A3"
    if analysis:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(IMPACT_COLS))}{len(analysis)+2}"

# ─── Template Checklist ───────────────────────────────────────

CHECKLIST_COLS = [
    ("Point de conformité", 40), ("Réf. source", 18), ("Statut", 14),
    ("Procédure", 28), ("Section couverte", 32), ("Écart / Remarque", 50),
    ("Action requise", 44), ("Criticité", 12), ("Confiance", 10),
]

def _write_checklist_sheet(ws, analysis, title_prefix):
    ws.title = "Checklist"
    ws.sheet_view.showGridLines = False
    _write_title(ws, f"{title_prefix} — ProcessMate", len(CHECKLIST_COLS))
    _write_headers(ws, CHECKLIST_COLS)

    for ri, item in enumerate(analysis, start=3):
        crit = item.get("criticality", "medium")
        cov  = item.get("coverage_status", "manquant")
        crit_bg, crit_fg = CRIT_COLORS.get(crit, CRIT_COLORS["medium"])
        cov_bg, cov_fg   = COVERAGE_COLORS.get(cov, COVERAGE_COLORS["manquant"])
        row_bg = WHITE if ri % 2 == 1 else GREY_BG

        first_action = (item.get("recommended_actions") or [{}])[0]
        proc_label = f"{item.get('procedure_ref','') or ''} {item.get('procedure_nom','') or ''}".strip()
        row_data = [
            item.get("source_element", ""),
            item.get("source_ref", ""),
            COVERAGE_LABELS.get(cov, cov),
            proc_label,
            item.get("procedure_section", ""),
            item.get("gap") or item.get("rationale", ""),
            first_action.get("description", "") if first_action else "",
            CRIT_LABELS.get(crit, crit),
            f"{round((item.get('confidence') or 0) * 100)}%",
        ]
        ws.row_dimensions[ri].height = _row_height(row_data, [c[1] for c in CHECKLIST_COLS])

        for ci, value in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.border = _border()
            if ci == 3:  # Statut
                c.font = _font(bold=True, size=9, color=cov_fg)
                c.fill = _fill(cov_bg); c.alignment = _align(wrap=False, h="center", v="center")
            elif ci == 8:  # Criticité
                c.font = _font(bold=True, size=9, color=crit_fg)
                c.fill = _fill(crit_bg); c.alignment = _align(wrap=False, h="center", v="center")
            else:
                c.font = _font(size=10); c.fill = _fill(row_bg); c.alignment = _align()

    ws.freeze_panes = "A3"
    if analysis:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(CHECKLIST_COLS))}{len(analysis)+2}"

# ─── Feuille Synthèse ─────────────────────────────────────────

def _write_summary_sheet(ws, artifact, summary, open_questions, analysis):
    ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    ws.merge_cells("A1:B1")
    ws["A1"].value = "SYNTHÈSE D'ANALYSE"
    ws["A1"].font = _font(bold=True, size=12, color="FFFFFFFF")
    ws["A1"].fill = _fill(BLUE_MID)
    ws["A1"].alignment = _align(wrap=False, h="center", v="center")
    ws.row_dimensions[1].height = 28

    covered  = sum(1 for a in analysis if a.get("coverage_status") == "couvert")
    partial  = sum(1 for a in analysis if a.get("coverage_status") == "partiel")
    missing  = sum(1 for a in analysis if a.get("coverage_status") == "manquant")
    critical = sum(1 for a in analysis if a.get("criticality") in ("critical", "high"))

    rows = [
        ("Type d'analyse",      artifact.get("intent_label", "")),
        ("Instruction",         artifact.get("instruction_summary", "")),
        ("Synthèse globale",    summary.get("global_assessment", "")),
        ("Sources identifiées", ", ".join(summary.get("sources_identified") or [])),
        ("Non impactées",       ", ".join(summary.get("procedures_not_impacted") or [])),
        ("Points analysés",     str(len(analysis))),
        ("Couverts ✓",          str(covered)),
        ("Partiels ~",          str(partial)),
        ("Manquants ✗",         str(missing)),
        ("Prioritaires",        str(critical)),
    ]
    for i, (label, value) in enumerate(rows, 2):
        ws.row_dimensions[i].height = max(18, min(14 * max(1, len(str(value)) // 90), 100))
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _font(bold=True, size=10, color="FF1E3A5F")
        lc.fill = _fill(BLUE_LIGHT if i % 2 == 0 else WHITE)
        lc.alignment = _align(wrap=False, h="left", v="top"); lc.border = _border()
        vc = ws.cell(row=i, column=2, value=value)
        vc.font = _font(size=10)
        vc.fill = _fill(BLUE_LIGHT if i % 2 == 0 else WHITE)
        vc.alignment = _align(wrap=True, h="left", v="top"); vc.border = _border()

    if open_questions:
        row_q = len(rows) + 3
        ws.merge_cells(f"A{row_q}:B{row_q}")
        hc = ws[f"A{row_q}"]
        hc.value = "QUESTIONS OUVERTES"
        hc.font = _font(bold=True, size=10, color="FFFFFFFF")
        hc.fill = _fill(BLUE_MID); hc.alignment = _align(wrap=False, h="center", v="center")
        ws.row_dimensions[row_q].height = 22

        for qi, q in enumerate(open_questions, row_q + 1):
            blocking = q.get("blocking", False)
            ws.row_dimensions[qi].height = max(18, min(14 * max(1, len(q.get("question","")) // 90), 100))
            lc = ws.cell(row=qi, column=1, value=f"{'🔴 Bloquant' if blocking else '🟡 À clarifier'} — {q.get('target','')}")
            lc.font = _font(bold=True, size=9, color="FFC0392B" if blocking else "FFB7800A")
            lc.fill = _fill("FFFDECEA" if blocking else "FFFFF8E1")
            lc.alignment = _align(wrap=True); lc.border = _border()
            vc = ws.cell(row=qi, column=2, value=q.get("question", ""))
            vc.font = _font(size=10)
            vc.fill = _fill("FFFDECEA" if blocking else "FFFFF8E1")
            vc.alignment = _align(wrap=True); vc.border = _border()

    ws.freeze_panes = "A2"

# ─── Feuille Journal ──────────────────────────────────────────

def _write_log_sheet(ws, analysis_log):
    ws.title = "Journal d'analyse"
    ws.sheet_view.showGridLines = False
    col_defs = [
        ("Procédure", 30), ("Sections examinées", 36),
        ("Résultats", 68), ("Points", 10), ("Justification", 54),
    ]
    _write_title(ws, "JOURNAL D'ANALYSE IA", len(col_defs))
    _write_headers(ws, col_defs)

    for ri, entry in enumerate(analysis_log, 3):
        n = entry.get("points_analyzed") or 0
        row_bg = "FFE8F8F5" if n > 0 else (WHITE if ri % 2 == 1 else GREY_BG)
        row_data = [
            entry.get("procedure_nom", ""),
            ", ".join(entry.get("examined_sections") or []),
            entry.get("findings", ""),
            n,
            entry.get("rationale", ""),
        ]
        ws.row_dimensions[ri].height = _row_height(row_data, [c[1] for c in col_defs])
        for ci, value in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.border = _border(); c.fill = _fill(row_bg)
            if ci == 4:
                c.font = _font(bold=True, size=11, color="FF1E8449" if n > 0 else "FF6C757D")
                c.alignment = _align(wrap=False, h="center", v="center")
            else:
                c.font = _font(size=10); c.alignment = _align()

    ws.freeze_panes = "A3"

# ─── Fonction principale ──────────────────────────────────────

def generate_analysis_excel(
    artifact: Dict[str, Any],
    analysis: List[Dict[str, Any]],
    summary: Dict[str, Any],
    analysis_log: List[Dict[str, Any]],
    open_questions: List[Dict[str, Any]],
    excel_template: str = "impact",
) -> bytes:
    wb = Workbook()
    ws_main = wb.active
    intent_label = artifact.get("intent_label") or "Analyse"

    if excel_template in ("checklist", "coverage_check"):
        _write_checklist_sheet(ws_main, analysis, intent_label)
    else:
        _write_impact_sheet(ws_main, analysis, intent_label)

    ws_summary = wb.create_sheet()
    _write_summary_sheet(ws_summary, artifact, summary, open_questions, analysis)

    if analysis_log:
        ws_log = wb.create_sheet()
        _write_log_sheet(ws_log, analysis_log)

    wb.properties.title = f"{intent_label} — ProcessMate"
    wb.properties.creator = "ProcessMate"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()