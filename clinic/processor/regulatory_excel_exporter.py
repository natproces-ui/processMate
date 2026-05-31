"""
Générateur Excel backend pour l'analyse d'impact réglementaire ProcessMate.
Utilise openpyxl pour une mise en forme professionnelle côté serveur.
"""

from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ─── Palette ──────────────────────────────────────────────────

BLUE_DARK   = "1E3A5F"   # header principal
BLUE_MID    = "2D6A9F"   # header secondaire
BLUE_LIGHT  = "D9EAF7"   # fond header colonnes
BLUE_ACCENT = "4A90D9"   # bordures accent
GREY_BG     = "F8F9FA"   # fond lignes paires
WHITE       = "FFFFFF"
BLACK       = "000000"

# Criticité
CRIT_COLORS = {
    "critical": ("FDECEA", "C0392B"),
    "high":     ("FEF3E2", "D35400"),
    "medium":   ("FFF8E1", "B7800A"),
    "low":      ("F0F4F8", "546E7A"),
}
CRIT_LABELS = {
    "critical": "Critique",
    "high":     "Élevé",
    "medium":   "Moyen",
    "low":      "Faible",
}
STATUS_LABELS = {
    "draft":     "Brouillon",
    "to_review": "À revoir",
    "validated": "Validé",
    "rejected":  "Rejeté",
    "converted": "Transformé",
}


# ─── Styles helpers ───────────────────────────────────────────

def _font(bold=False, size=10, color=BLACK, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _border(style="thin", color="D0D7DE") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(wrap=True, h="left", v="top") -> Alignment:
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)


def _set_cell(ws, row: int, col: int, value, **kwargs):
    cell = ws.cell(row=row, column=col, value=value)
    if "font" in kwargs:      cell.font = kwargs["font"]
    if "fill" in kwargs:      cell.fill = kwargs["fill"]
    if "alignment" in kwargs: cell.alignment = kwargs["alignment"]
    if "border" in kwargs:    cell.border = kwargs["border"]
    return cell


# ─── Feuille Impacts ──────────────────────────────────────────

IMPACT_COLUMNS = [
    ("Activité",              28),
    ("Thème",                 32),
    ("Changement introduit",  58),
    ("Dép. BAM",              12),
    ("Impact Métier",         52),
    ("Impact SI",             52),
    ("Systèmes impactés",     22),
    ("Actions recommandées",  62),
    ("Commentaire SI",        36),
    ("Procédure",             30),
    ("Criticité",             12),
    ("Statut",                12),
    ("Référence loi",         26),
    ("Confiance IA",          12),
]


def _write_impacts_sheet(ws, impacts: List[Dict[str, Any]]) -> None:
    ws.title = "Impacts"
    ws.sheet_view.showGridLines = False

    # ── Titre de la feuille ──
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = "ANALYSE D'IMPACT RÉGLEMENTAIRE — ProcessMate"
    title_cell.font = _font(bold=True, size=13, color=WHITE)
    title_cell.fill = _fill(BLUE_DARK)
    title_cell.alignment = _align(wrap=False, h="center", v="center")
    ws.row_dimensions[1].height = 30

    # ── En-têtes colonnes ──
    ws.row_dimensions[2].height = 32
    header_border = _border("medium", BLUE_ACCENT)
    for ci, (label, width) in enumerate(IMPACT_COLUMNS, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.font = _font(bold=True, size=10, color=BLUE_DARK)
        cell.fill = _fill(BLUE_LIGHT)
        cell.alignment = _align(wrap=True, h="center", v="center")
        cell.border = header_border
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── Données ──
    for ri, impact in enumerate(impacts, start=3):
        meta = impact.get("metadata") or {}
        raw = meta.get("raw") or meta
        activity  = raw.get("activity") or impact.get("category") or ""
        si_comment= raw.get("si_comment") or ""
        ext_dep   = impact.get("external_dependency") or raw.get("external_dependency") or ""
        crit      = impact.get("criticality", "medium")
        crit_bg, crit_fg = CRIT_COLORS.get(crit, CRIT_COLORS["medium"])
        row_bg    = WHITE if ri % 2 == 1 else GREY_BG
        std_fill  = _fill(row_bg)
        std_border= _border("thin", "D0D7DE")
        std_align = _align(wrap=True, h="left", v="top")

        actions_text = "\n".join(
            f"• [{a.get('priority','').upper()}] {a.get('title','')} — {a.get('description','')}"
            for a in (impact.get("recommended_actions") or [])
        )

        row_data = [
            activity,
            impact.get("theme", ""),
            impact.get("regulatory_change", ""),
            "OUI" if ext_dep else "NON",
            impact.get("business_impact", ""),
            impact.get("si_impact", ""),
            "\n".join(impact.get("impacted_systems") or []),
            actions_text,
            si_comment,
            f"{impact.get('procedure_ref','') or ''} {impact.get('procedure_nom','') or ''}".strip(),
            CRIT_LABELS.get(crit, crit),
            STATUS_LABELS.get(impact.get("status",""), impact.get("status","")),
            impact.get("law_reference", ""),
            f"{round((impact.get('confidence') or 0) * 100)}%",
        ]

        # Estimer hauteur selon contenu
        max_lines = 1
        for col_idx, value in enumerate(row_data):
            col_width = IMPACT_COLUMNS[col_idx][1]
            text = str(value or "")
            lines = text.count("\n") + 1
            wrap_lines = max(1, len(text) // max(col_width, 10))
            max_lines = max(max_lines, lines, wrap_lines)
        ws.row_dimensions[ri].height = min(max(max_lines * 14, 18), 220)

        for ci, value in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = std_border

            # Colonne criticité : badge coloré
            if ci == 11:
                cell.font = _font(bold=True, size=9, color=crit_fg)
                cell.fill = _fill(crit_bg)
                cell.alignment = _align(wrap=False, h="center", v="center")
            # Colonne dépendance BAM
            elif ci == 4:
                is_oui = bool(ext_dep)
                cell.font = _font(bold=True, size=9, color="C0392B" if is_oui else "27AE60")
                cell.fill = _fill("FDECEA" if is_oui else "EAFAF1")
                cell.alignment = _align(wrap=False, h="center", v="center")
            # Colonne confiance IA
            elif ci == 14:
                cell.font = _font(bold=False, size=9, color="2D6A9F")
                cell.fill = std_fill
                cell.alignment = _align(wrap=False, h="center", v="center")
            # Colonne statut
            elif ci == 12:
                status = impact.get("status", "")
                s_colors = {
                    "validated": ("EAFAF1", "27AE60"),
                    "rejected":  ("FDECEA", "C0392B"),
                    "to_review": ("FFF8E1", "B7800A"),
                    "converted": ("EBF5FB", "2980B9"),
                }
                sbg, sfg = s_colors.get(status, (row_bg, "546E7A"))
                cell.font = _font(bold=False, size=9, color=sfg)
                cell.fill = _fill(sbg)
                cell.alignment = _align(wrap=False, h="center", v="center")
            else:
                cell.font = _font(size=10)
                cell.fill = std_fill
                cell.alignment = std_align

    # Figer les 2 premières lignes
    ws.freeze_panes = "A3"

    # Filtre automatique
    if impacts:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(IMPACT_COLUMNS))}{len(impacts) + 2}"


# ─── Feuille Diagnostic ───────────────────────────────────────

def _write_diagnostic_sheet(ws, campaign: Dict[str, Any], last_analysis: Optional[Dict]) -> None:
    ws.title = "Diagnostic"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110

    # Titre
    ws.merge_cells("A1:B1")
    ws["A1"].value = "DIAGNOSTIC D'ANALYSE"
    ws["A1"].font = _font(bold=True, size=12, color=WHITE)
    ws["A1"].fill = _fill(BLUE_MID)
    ws["A1"].alignment = _align(wrap=False, h="center", v="center")
    ws.row_dimensions[1].height = 28

    summary = (last_analysis or {}).get("summary") or {}
    procedures_analyzed = (last_analysis or {}).get("procedures_analyzed") or []

    rows = [
        ("Campagne",              campaign.get("title", "")),
        ("Statut",                campaign.get("status", "")),
        ("Source",                campaign.get("source_filename") or campaign.get("source_type") or ""),
        ("Sujet réglementaire",   summary.get("regulatory_subject", "")),
        ("Synthèse globale",      summary.get("global_assessment", "")),
        ("Modèle IA",             (last_analysis or {}).get("model_used", "")),
        ("Procédures analysées",  ", ".join(f"{p.get('nom','')} ({p.get('steps_count',0)} étapes)" for p in procedures_analyzed)),
        ("Impacts bruts",         str((last_analysis or {}).get("raw_impacts_count", ""))),
        ("Impacts créés",         str((last_analysis or {}).get("impacts_count", ""))),
        ("Non rattachés",         str(len((last_analysis or {}).get("unresolved_impacts") or []))),
        ("Dépendances",           ", ".join(summary.get("dependencies") or [])),
        ("Non impactées",         ", ".join(summary.get("procedures_not_impacted") or [])),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = max(18, min(14 * (str(value).count("\n") + max(1, len(str(value)) // 100)), 120))
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _font(bold=True, size=10, color=BLUE_DARK)
        lc.fill = _fill(BLUE_LIGHT if i % 2 == 0 else WHITE)
        lc.alignment = _align(wrap=False, h="left", v="top")
        lc.border = _border("thin", "D0D7DE")

        vc = ws.cell(row=i, column=2, value=value)
        vc.font = _font(size=10)
        vc.fill = _fill(BLUE_LIGHT if i % 2 == 0 else WHITE)
        vc.alignment = _align(wrap=True, h="left", v="top")
        vc.border = _border("thin", "D0D7DE")

    # Questions ouvertes
    questions = (last_analysis or {}).get("open_questions") or []
    if questions:
        row_q = len(rows) + 3
        ws.merge_cells(f"A{row_q}:B{row_q}")
        hc = ws[f"A{row_q}"]
        hc.value = "QUESTIONS OUVERTES"
        hc.font = _font(bold=True, size=10, color=WHITE)
        hc.fill = _fill(BLUE_MID)
        hc.alignment = _align(wrap=False, h="center", v="center")
        ws.row_dimensions[row_q].height = 22

        for qi, q in enumerate(questions, start=row_q + 1):
            blocking = q.get("blocking", False)
            label = f"{'🔴 Bloquant' if blocking else '🟡 À clarifier'} — {q.get('target','')}"
            ws.row_dimensions[qi].height = max(18, min(14 * (str(q.get('question','')).count("\n") + max(1, len(str(q.get('question',''))) // 100)), 100))

            lc = ws.cell(row=qi, column=1, value=label)
            lc.font = _font(bold=True, size=9, color="C0392B" if blocking else "B7800A")
            lc.fill = _fill("FDECEA" if blocking else "FFF8E1")
            lc.alignment = _align(wrap=True, h="left", v="top")
            lc.border = _border("thin", "D0D7DE")

            vc = ws.cell(row=qi, column=2, value=q.get("question", ""))
            vc.font = _font(size=10)
            vc.fill = _fill("FDECEA" if blocking else "FFF8E1")
            vc.alignment = _align(wrap=True, h="left", v="top")
            vc.border = _border("thin", "D0D7DE")

    ws.freeze_panes = "A2"


# ─── Feuille Journal d'analyse ────────────────────────────────

def _write_journal_sheet(ws, analysis_log: List[Dict[str, Any]]) -> None:
    ws.title = "Journal d'analyse"
    ws.sheet_view.showGridLines = False

    col_defs = [
        ("Procédure",          30),
        ("Sections examinées", 36),
        ("Résultats",          70),
        ("Impacts créés",      14),
        ("Justification",      56),
    ]
    for ci, (_, w) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Titre
    ws.merge_cells(f"A1:{get_column_letter(len(col_defs))}1")
    ws["A1"].value = "JOURNAL D'ANALYSE IA"
    ws["A1"].font = _font(bold=True, size=12, color=WHITE)
    ws["A1"].fill = _fill(BLUE_MID)
    ws["A1"].alignment = _align(wrap=False, h="center", v="center")
    ws.row_dimensions[1].height = 28

    # Headers
    ws.row_dimensions[2].height = 28
    for ci, (label, _) in enumerate(col_defs, 1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.font = _font(bold=True, size=10, color=BLUE_DARK)
        cell.fill = _fill(BLUE_LIGHT)
        cell.alignment = _align(wrap=True, h="center", v="center")
        cell.border = _border("medium", BLUE_ACCENT)

    for ri, entry in enumerate(analysis_log, start=3):
        impacts_count = entry.get("impacts_created") or 0
        row_bg = "EAFAF1" if impacts_count > 0 else WHITE if ri % 2 == 1 else GREY_BG
        sections_text = ", ".join(entry.get("examined_sections") or [])
        row_data = [
            entry.get("procedure_nom", ""),
            sections_text,
            entry.get("findings", ""),
            impacts_count,
            entry.get("rationale", ""),
        ]
        max_lines = 1
        for col_idx, value in enumerate(row_data):
            col_width = col_defs[col_idx][1]
            text = str(value or "")
            lines = text.count("\n") + 1
            wrap_lines = max(1, len(text) // max(col_width, 10))
            max_lines = max(max_lines, lines, wrap_lines)
        ws.row_dimensions[ri].height = min(max(max_lines * 14, 18), 180)

        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = _border("thin", "D0D7DE")
            cell.fill = _fill(row_bg)
            if ci == 4:
                cell.font = _font(bold=True, size=11, color="27AE60" if impacts_count > 0 else "546E7A")
                cell.alignment = _align(wrap=False, h="center", v="center")
            else:
                cell.font = _font(size=10)
                cell.alignment = _align(wrap=True, h="left", v="top")

    ws.freeze_panes = "A3"


# ─── Fonction principale ──────────────────────────────────────

def generate_impact_excel(
    campaign: Dict[str, Any],
    impacts: List[Dict[str, Any]],
    last_analysis: Optional[Dict[str, Any]] = None,
    analysis_log: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    """
    Génère le fichier Excel d'analyse d'impact et retourne les bytes.
    Ordre des feuilles : Impacts → Diagnostic → Journal d'analyse
    """
    wb = Workbook()

    # Feuille Impacts (feuille active par défaut)
    ws_impacts = wb.active
    _write_impacts_sheet(ws_impacts, impacts)

    # Feuille Diagnostic
    ws_diag = wb.create_sheet()
    _write_diagnostic_sheet(ws_diag, campaign, last_analysis)

    # Feuille Journal (si données disponibles)
    if analysis_log:
        ws_log = wb.create_sheet()
        _write_journal_sheet(ws_log, analysis_log)

    # Propriétés du classeur
    wb.properties.title = f"Analyse Impact — {campaign.get('title', 'ProcessMate')}"
    wb.properties.creator = "ProcessMate"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()